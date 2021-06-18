from os.path import join
import cv2
import os
import sys
import traceback
from openpyxl import Workbook
from itertools import product
import numpy as np
from collections import Counter
import matplotlib.pyplot as plt
from matplotlib.gridspec import GridSpec
import matplotlib as mpl

mpl.rcParams['font.family'] = 'Times New Roman'
mpl.rcParams['mathtext.fontset'] = 'cm'

sys.path.append('..')
from utils import option_choice, init, select_data_file, load_session, grid
from utils.io import pd_to_sheet, player_round_mat_to_sheet, OTreeSessionData

DEBUG = False
DFT_W, DFT_H = 7, 7
DATA_DIR = 'data'
OUTPUT_DIR = 'output'
APP_NAME = 'pd_node_edge'
I_ROUND = 0
I_ID = 1

global_vars = dict()


def stats_strategy_response(data: OTreeSessionData, ws=None):
    """If `ws` is given, save to excel rows."""

    rounds = [data.get_round(i + 1) for i in range(data.num_rounds())]
    rnd_stats = {}
    for rnd, (pr, ne) in enumerate(zip(rounds[:-1], rounds[1:]), 2):
        stats = {k1: {k2: 0 for k2 in product(('Node', 'Edge'), repeat=2)}
                 for k1 in product(list('CD'), repeat=2)}
        print(f'\rCalculating strategy response (round {rnd}) ... ', end='')
        for pid in range(1, ne.num_players() + 1):
            nep = ne.loc[pid]
            prp = pr.loc[pid]
            stats[(nep.player.choice_L, prp.player.choice_nei_L)][(nep.player.type, prp.player.type_L)] += 1
            stats[(nep.player.choice_U, prp.player.choice_nei_U)][(nep.player.type, prp.player.type_U)] += 1
            stats[(nep.player.choice_R, prp.player.choice_nei_R)][(nep.player.type, prp.player.type_R)] += 1
            stats[(nep.player.choice_D, prp.player.choice_nei_D)][(nep.player.type, prp.player.type_D)] += 1
            if ws:
                ws.append((rnd, pid, nep.player.type,
                           nep.player.choice_L, nep.player.type_L, prp.player.choice_nei_L,
                           nep.player.choice_U, nep.player.type_U, prp.player.choice_nei_U,
                           nep.player.choice_R, nep.player.type_R, prp.player.choice_nei_R,
                           nep.player.choice_D, nep.player.type_D, prp.player.choice_nei_D))
            rnd_stats[rnd] = stats
    print('Done')
    global_vars['rnd_stats'] = rnd_stats
    return rnd_stats


def save_neighbor_info(session_code, data: OTreeSessionData, w, h):
    """Save a excel with all neighbors' information."""

    file_name = join(OUTPUT_DIR, f'{APP_NAME}_neighbors_{session_code}.xlsx')
    wb = Workbook()

    ws0 = wb.active
    ws0.title = "All"
    pd_to_sheet(ws0, data, 'type payoff pid_L pid_U pid_R pid_D '
                           'choice_L choice_U choice_R choice_D '
                           'type_L type_U type_R type_D '
                           'choice_nei_L choice_nei_U choice_nei_R choice_nei_D'.split(' '))
    player_round_mat_to_sheet(wb.create_sheet(title='payoff'), data.player.payoff)
    player_round_mat_to_sheet(wb.create_sheet(title='choice_L'), data.player.choice_L)
    player_round_mat_to_sheet(wb.create_sheet(title='choice_U'), data.player.choice_U)
    player_round_mat_to_sheet(wb.create_sheet(title='choice_R'), data.player.choice_R)
    player_round_mat_to_sheet(wb.create_sheet(title='choice_D'), data.player.choice_D)
    player_round_mat_to_sheet(wb.create_sheet(title='choice_nei_L'), data.player.choice_nei_L)
    player_round_mat_to_sheet(wb.create_sheet(title='choice_nei_U'), data.player.choice_nei_U)
    player_round_mat_to_sheet(wb.create_sheet(title='choice_nei_R'), data.player.choice_nei_R)
    player_round_mat_to_sheet(wb.create_sheet(title='choice_nei_D'), data.player.choice_nei_D)
    wb.save(file_name)


def save_strategy_response(session_code, data: OTreeSessionData, w, h):
    """Save a excel with strategy response."""

    file_name = join(OUTPUT_DIR, f'{APP_NAME}_response_{session_code}.xlsx')
    wb = Workbook()

    ws0 = wb.active
    ws0.title = "All"
    ws0.append('Round player type '
               'choice_L type_L nei_L_last_choice '
               'choice_U type_U nei_U_last_choice '
               'choice_R type_R nei_R_last_choice '
               'choice_D type_D nei_D_last_choice'.split(' '))

    rnd_stats = stats_strategy_response(data, ws0)
    ws1 = wb.create_sheet(title='Stats')
    ws1.append('Round CC CD DC DD'.split() +
               list(map(lambda x: x[1] + x[0], product('CC CD DC DD'.split(), 'nn nl ln ll'.split()))))
    for rnd, stats in rnd_stats.items():
        ws1.append([rnd] + list(map(lambda s: sum(s.values()), stats.values())) +
                   sum(map(lambda s: list(s.values()), stats.values()), []))
    print('Done')

    wb.save(file_name)


STRATEGY_RESPONSE_COLOR_MAP = [
    '#733d11', '#a65819', '#d97321', '#ff8826',
    '#5a7328', '#82a63a', '#aad94c', '#c8ff59',
    '#116473', '#1991a6', '#21bdd9', '#26deff',
    '#5c1773', '#8521a6', '#ad2bd9', '#cc33ff',
]


def save_strategy_response_barplot(session_code, data: OTreeSessionData, w, h):
    """Save a bar plot of strategy response statistics."""
    file_name = join(OUTPUT_DIR, f'{APP_NAME}_response_stats_{session_code}.pdf')
    rnd_stats = global_vars.get('rnd_stats', None)
    if rnd_stats is None: rnd_stats = stats_strategy_response(data)

    ratio = []
    rounds = list(range(2, data.num_rounds() + 1))
    for rnd, stats in rnd_stats.items():
        values = sum(map(lambda s: list(s.values()), stats.values()), [])
        ratio.append([v / sum(values) for v in values])
    ratio = np.array(ratio).T
    fig, ax = plt.subplots()
    ax.bar(rounds, ratio[0])
    start = np.zeros(ratio[0].shape)
    for i, r in enumerate(ratio):
        if sum(r) < 1e-6: continue
        ax.bar(rounds, r, bottom=start, width=1, color=STRATEGY_RESPONSE_COLOR_MAP[i])
        start += r
    ax.axis([min(rounds) - 0.5, max(rounds) + 0.5, 0, 1])
    ax.set_xlabel('Rounds')
    ax.set_ylabel('Ratio')
    fig.savefig(file_name)
    fig.show()


VIDEO_COLOR_MAP = dict(
    Node=dict(C='#EA7462', D='#6A8FE6'),
    Edge=dict(C='#B32A15', D='#1C47AD')
)

BAR_COLOR_MAP = ['#2858ba', '#414890', '#663862', '#9b3641', '#ca362c']


def round_plot(file_name, dict_data, rnd, max_rnd, w, h,
               font_suptitle=20, font_title=10, stats_data=None, hist_CD=None):
    plt.close('all')
    fig = plt.Figure(dpi=300)
    #     gs = GridSpec(8, 14, wspace=2, hspace=5)
    #     ax_pattern = fig.add_subplot(gs[:, :8])
    #     ax_evolution = fig.add_subplot(gs[:3, 8:])
    #     ax_payoff = fig.add_subplot(gs[3:-2, 8:-3])
    #     ax_choice_stats = fig.add_subplot(gs[3:-2, -3:])
    #     ax_response = fig.add_subplot(gs[-2:, 8:])
    gs = GridSpec(16, 30)
    ax_pattern = fig.add_subplot(gs[:, :16])
    ax_evolution = fig.add_subplot(gs[2:5, 17:])
    ax_payoff = fig.add_subplot(gs[6:-5, 17:-7])
    ax_choice_stats = fig.add_subplot(gs[6:-5, -6:])
    ax_response = fig.add_subplot(gs[-3:-2, 17:])

    def _plot_pattern(ax, data):
        def _fill_xy(i, j, _type):
            points = [[i - 0.5, j - 0.5], [i - 0.5, j + 0.5], [i + 0.5, j + 0.5], [i + 0.5, j - 0.5],
                      [i - 0.5, j - 0.5]]
            if _type == 'Node':
                return np.array(points[:-1])
            else:
                return dict(
                    list(zip(list('LURD'), [np.array([[i, j]] + [p1, p2]) for p1, p2 in zip(points[:-1], points[1:])])))

        def _plot_player(i, j, d):
            if d['type'] == 'Node':
                xy = _fill_xy(i, j, d['type'])
                ax.fill(xy[:, 0], xy[:, 1], VIDEO_COLOR_MAP[d['type']][d['L']])
            else:
                for di, xy in _fill_xy(i, j, d['type']).items():
                    ax.fill(xy[:, 0], xy[:, 1], VIDEO_COLOR_MAP[d['type']][d[di]])

        for pid, p_data in data.items():
            pi = (pid - 1) // w + 1
            pj = (pid - 1) % w + 1
            _plot_player(pi, pj, p_data)
        ax.set_aspect('equal')
        ax.set_title('Pattern', fontsize=font_title)
        ax.set_xlim([0.5, w + 0.5])
        ax.set_ylim([0.5, h + 0.5])
        ax.get_xaxis().set_visible(False)
        ax.get_yaxis().set_visible(False)
        # ax.axis('off')

    def _plot_evolution(ax, hist_data):
        if hist_data:
            Cs = np.array([cnt['C'] for cnt in hist_data])
            Ds = np.array([cnt['D'] for cnt in hist_data])
            CDs = Cs + Ds
            rounds = np.array(list(range(1, len(hist_data) + 1)))
            ax.plot(rounds, Cs / CDs, '-', color=BAR_COLOR_MAP[-1], label='C')
            ax.plot(rounds, Ds / CDs, '-', color=BAR_COLOR_MAP[0], label='D')
            ax.legend(ncol=2, fontsize=6, markerscale=0.8)
        ax.set_yticks([0, 1])
        ax.set_xlim([1, max_rnd])
        ax.set_ylim([0, 1])
        ax.set_title('Time steps', fontsize=font_title)
        ax.spines['right'].set_visible(False)
        ax.spines['top'].set_visible(False)

    def _plot_payoff_hist(ax, data):
        cm = plt.cm.get_cmap("Blues")
        _, bins, patches = ax.hist([v['payoff'] for v in data.values()], density=True, range=(-4, 24), bins=12)
        for idx, patch in enumerate(patches):
            patch.set_facecolor(cm((idx + 3) / (len(patches)+2)))
        ax.set_xticks([-4, 10, 24])
        ax.set_xlim([-4.5, 24.5])
        ax.set_ylabel('Payoff', fontsize=font_title, labelpad=0.1)
        ax.get_yaxis().set_ticks([])
        ax.spines['left'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['top'].set_visible(False)

    def _plot_choice_stats(ax, data):
        cnt = Counter([''.join([v[_di] for _di in 'LURD']).count('C') for v in data.values()])
        if cnt[1] + cnt[2] + cnt[3] == 0:
            ax.pie([cnt[0], cnt[4]], counterclock=False, startangle=90,
                   labels=['0', '1'], colors=[BAR_COLOR_MAP[0], BAR_COLOR_MAP[-1]],
                   textprops=dict(fontsize=6, color='#FFFFFF'), labeldistance=0.5)
            pass
        else:
            ax.pie([cnt[c] for c in range(5)], counterclock=False, startangle=90,
                   labels=['0', '1/4', '1/2', '3/4', '1'], colors=BAR_COLOR_MAP,
                   textprops=dict(fontsize=6, color='#FFFFFF'), labeldistance=0.5)
        ax.set_ylabel('Choices', fontsize=font_title, labelpad=0.1)
        ax.spines['right'].set_visible(False)
        ax.spines['top'].set_visible(False)

    def _plot_response(ax, stats_list):
        if stats_list:
            stats_total = sum(stats_list)
            ratio = [v / stats_total for v in stats_list]
            ax.barh([1], ratio[0])
            start = 0
            for i, r in enumerate(ratio):
                if r < 1e-6: continue
                ax.barh([1], r, left=start, height=1, color=STRATEGY_RESPONSE_COLOR_MAP[i])
                start += r
        ax.set_title('Response', fontsize=font_title, pad=5)
        ax.get_xaxis().set_visible(False)
        ax.get_yaxis().set_visible(False)
        ax.set_xlim([0, 1])
        ax.set_ylim([0.5, 1.5])

    _plot_pattern(ax_pattern, dict_data)
    _plot_evolution(ax_evolution, hist_CD)
    _plot_payoff_hist(ax_payoff, dict_data)
    _plot_choice_stats(ax_choice_stats, dict_data)
    _plot_response(ax_response, sum([list(v.values()) for v in stats_data.values()], []) if stats_data else None)

    fig.suptitle(f'Round {rnd}', fontsize=font_suptitle, y=0.9)
    fig.subplots_adjust(left=0.05, bottom=0.01,
                        right=0.95, top=0.9,
                        wspace=1.2, hspace=0)
    # fig.tight_layout()
    fig.savefig(file_name)


def release_video(file_name, image_name_list, fps=24, video_type='mp4'):
    assert len(image_name_list) > 0, "No images found for rendering."
    image = cv2.imread(image_name_list[0])
    height, width, _ = image.shape
    codecs = {
        'mp4': cv2.VideoWriter_fourcc(*'MP4V'),
        'avi': cv2.VideoWriter_fourcc(*'DIVX'),
    }
    assert video_type.lower() in codecs, 'Error video type.'

    video = cv2.VideoWriter(file_name, codecs[video_type], float(fps),
                            (width, height))
    for i, image_name in enumerate(image_name_list, 1):
        print(f'\rRelease video [Round {i}] ... ', end='')
        for _ in range(int(fps)):
            image = cv2.imread(image_name)
            video.write(image)
    video.release()
    cv2.destroyAllWindows()
    print('Done')


def render_video(session_code, data: OTreeSessionData, w, h):
    """Save a video for this session."""
    dir_name = join(OUTPUT_DIR, f'{APP_NAME}_render_{session_code}')
    video_name = join(dir_name, f'{session_code}_output.mp4')
    rnd_stats = global_vars.get('rnd_stats', None)
    if rnd_stats is None: rnd_stats = stats_strategy_response(data)
    if not os.path.exists(dir_name): os.mkdir(dir_name)
    image_name_list = []
    cnt_CD = []
    for rnd in range(1, data.num_rounds() + 1):
        fig_name = join(dir_name, f'Round_{rnd}.png')
        print(f'\rRender image for round {rnd} ... ', end='')
        round_data = data.get_round(rnd)
        round_plot_data = dict()
        for pid in range(1, round_data.num_players() + 1):
            d = round_data.loc[pid]
            round_plot_data[pid] = dict(type=d.player.type, L=d.player.choice_L, U=d.player.choice_U,
                                        R=d.player.choice_R, D=d.player.choice_D, payoff=d.player.payoff)
        cnt = Counter(''.join([''.join([v[_di] for _di in 'LURD']) for v in round_plot_data.values()]))
        cnt_CD.append(cnt)
        round_plot(fig_name, round_plot_data, rnd, data.num_rounds(),
                   w, h, stats_data=rnd_stats.get(rnd, None), hist_CD=cnt_CD)
        image_name_list.append(fig_name)
    print('Done')
    release_video(video_name, image_name_list, video_type='mp4')


def fill_edge_choice(data: OTreeSessionData):
    nodes = data.player.node_choice.notna()
    data.loc[nodes,
             [('player', 'choice_L'),
              ('player', 'choice_U'),
              ('player', 'choice_R'),
              ('player', 'choice_D')
              ]] = data.loc[nodes, ('player', 'node_choice')]
    data.loc[:, ('player', 'type')] = np.where(nodes, 'Node', 'Edge')
    return data[data.player.choice_L.notna()]


def calculate_data(data: OTreeSessionData, w, h):
    nei = grid.neighbor_mapping(w, h)
    data = data.reindex(columns=data.columns.tolist()
                                + list(map(lambda x: ('player', x),
                                           'pid_L pid_U pid_R pid_D '
                                           'type_L type_U type_R type_D '
                                           'choice_nei_L choice_nei_U choice_nei_R choice_nei_D'.split(' '))))
    for idx, row in data.iterrows():
        rnd, pid = idx
        print(f'\rProducing neighbor info (round {rnd}, player {pid}) ... ', end='')
        data.loc[idx, ('player', 'pid_L')] = nei[pid]['L']
        data.loc[idx, ('player', 'pid_U')] = nei[pid]['U']
        data.loc[idx, ('player', 'pid_R')] = nei[pid]['R']
        data.loc[idx, ('player', 'pid_D')] = nei[pid]['D']
        data.loc[idx, ('player', 'type_L')] = data.loc[(rnd, nei[pid]['L']), ('player', 'type')]
        data.loc[idx, ('player', 'type_U')] = data.loc[(rnd, nei[pid]['U']), ('player', 'type')]
        data.loc[idx, ('player', 'type_R')] = data.loc[(rnd, nei[pid]['R']), ('player', 'type')]
        data.loc[idx, ('player', 'type_D')] = data.loc[(rnd, nei[pid]['D']), ('player', 'type')]
        data.loc[idx, ('player', 'choice_nei_L')] = data.loc[(rnd, nei[pid]['L']), ('player', 'choice_R')]
        data.loc[idx, ('player', 'choice_nei_U')] = data.loc[(rnd, nei[pid]['U']), ('player', 'choice_D')]
        data.loc[idx, ('player', 'choice_nei_R')] = data.loc[(rnd, nei[pid]['R']), ('player', 'choice_L')]
        data.loc[idx, ('player', 'choice_nei_D')] = data.loc[(rnd, nei[pid]['D']), ('player', 'choice_U')]
    print('Done')
    return data


OPERATION = [
    save_neighbor_info,
    save_strategy_response,
    save_strategy_response_barplot,
    render_video,
]


def run():
    init(DATA_DIR, OUTPUT_DIR)
    file_name = select_data_file(DATA_DIR)
    session_code, data = load_session(file_name)
    data = data.rename(
        columns=dict(left_edge='choice_L', up_edge='choice_U', right_edge='choice_R', down_edge='choice_D'))
    num_players = len(data.iloc[data.index.get_level_values('round_number') == 1])
    w, h = grid.ask_W_H(num_players) if not DEBUG else (DFT_W, DFT_H)
    data = fill_edge_choice(data)
    data = calculate_data(data, w, h)
    if not DEBUG:
        while True:
            OPERATION[option_choice(list(map(lambda x: x.__doc__, OPERATION)), "Choose one operation:")](
                session_code, data, w, h)
    else:
        # save_neighbor_info(session_code, data, w, h)
        # save_strategies_response(session_code, data, w, h)
        save_strategy_response_barplot(session_code, data, w, h)
        pass


if __name__ == '__main__':
    try:
        run()
    except Exception:
        traceback.print_exception(*sys.exc_info())
        if not DEBUG:
            os.system("pause")
