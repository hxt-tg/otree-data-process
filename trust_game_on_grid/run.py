from os.path import join
import os
import sys
from openpyxl import Workbook
import traceback

sys.path.append('..')
from utils import option_choice, init, select_data_file, load_session
from utils.io import pd_to_sheet, player_round_mat_to_sheet, OTreeSessionData

DEBUG = False
DATA_DIR = 'data'
OUTPUT_DIR = 'output'
APP_NAME = 'trust_game_on_grid'


def save_player_round(session_code, data: OTreeSessionData):
    """Save player-rounds matrix."""

    file_name = join(OUTPUT_DIR, f'{APP_NAME}_{session_code}.xlsx')
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "All"

    pd_to_sheet(ws0, data, 'send_T return_T receive_send_T receive_return_T payoff is_node_player'.split(' '))
    player_round_mat_to_sheet(wb.create_sheet(title='send_T'), data.player.send_T)
    player_round_mat_to_sheet(wb.create_sheet(title='return_T'), data.player.return_T)
    player_round_mat_to_sheet(wb.create_sheet(title='receive_send_T'), data.player.receive_send_T)
    player_round_mat_to_sheet(wb.create_sheet(title='receive_return_T'), data.player.receive_return_T)
    player_round_mat_to_sheet(wb.create_sheet(title='payoff'), data.player.payoff)
    wb.save(file_name)


OPERATION = [
    save_player_round
]


def run():
    init(DATA_DIR, OUTPUT_DIR)
    file_name = select_data_file(DATA_DIR)
    session_code, data = load_session(file_name)
    if not DEBUG:
        while True:
            OPERATION[option_choice(list(map(lambda x: x.__doc__, OPERATION)), "Choose one operation:")](
                session_code, data)
    else:
        save_player_round(session_code, data)


if __name__ == '__main__':
    try:
        run()
    except Exception:
        traceback.print_exception(*sys.exc_info())
        if not DEBUG:
            os.system("pause")
