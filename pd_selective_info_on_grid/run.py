from os.path import join
import os
import sys
import traceback
from openpyxl import Workbook
from itertools import product
import numpy as np

sys.path.append('..')
from utils import option_choice, init, select_data_file, load_session, grid
from utils.io import pd_to_sheet, player_round_mat_to_sheet, OTreeSessionData

DEBUG = False
DFT_W, DFT_H = 7, 7
DATA_DIR = 'data'
OUTPUT_DIR = 'output'
APP_NAME = 'pd_node_edge'
I_ROUND = 0
I_ID = 1

global_vars = dict()


def stats_strategy_response(data: OTreeSessionData, ws=None):
    """If `ws` is given, save to excel rows."""

    rounds = [data.get_round(i + 1) for i in range(data.num_rounds())]
    rnd_stats = {}
    for rnd, (pr, ne) in enumerate(zip(rounds[:-1], rounds[1:]), 2):
        stats = {k1: {k2: 0 for k2 in product(('All', 'Highest'), repeat=2)}
                 for k1 in product(list('CD'), repeat=2)}
        print(f'\rCalculating strategy response (round {rnd}) ...    ', end='')
        for pid in range(1, ne.num_players() + 1):
            nep = ne.loc[pid]
            prp = pr.loc[pid]
            stats[(nep.player.choice_L, prp.player.choice_nei_L)][(nep.player.info, prp.player.info_L)] += 1
            stats[(nep.player.choice_U, prp.player.choice_nei_U)][(nep.player.info, prp.player.info_U)] += 1
            stats[(nep.player.choice_R, prp.player.choice_nei_R)][(nep.player.info, prp.player.info_R)] += 1
            stats[(nep.player.choice_D, prp.player.choice_nei_D)][(nep.player.info, prp.player.info_D)] += 1
            if ws:
                ws.append((rnd, pid, nep.player.info,
                           nep.player.choice_L, nep.player.info_L, prp.player.choice_nei_L,
                           nep.player.choice_U, nep.player.info_U, prp.player.choice_nei_U,
                           nep.player.choice_R, nep.player.info_R, prp.player.choice_nei_R,
                           nep.player.choice_D, nep.player.info_D, prp.player.choice_nei_D))
            rnd_stats[rnd] = stats
    global_vars['rnd_stats'] = rnd_stats
    return rnd_stats


def save_neighbor_info(session_code, data: OTreeSessionData, w, h):
    """Save a excel with all neighbors' information."""

    file_name = join(OUTPUT_DIR, f'{APP_NAME}_neighbors_{session_code}.xlsx')
    wb = Workbook()

    ws0 = wb.active
    ws0.title = "All"
    pd_to_sheet(ws0, data, 'info payoff pid_L pid_U pid_R pid_D '
                           'choice_L choice_U choice_R choice_D '
                           'info_L info_U info_R info_D '
                           'choice_nei_L choice_nei_U choice_nei_R choice_nei_D'.split(' '))
    player_round_mat_to_sheet(wb.create_sheet(title='payoff'), data.player.payoff)
    player_round_mat_to_sheet(wb.create_sheet(title='choice_L'), data.player.choice_L)
    player_round_mat_to_sheet(wb.create_sheet(title='choice_U'), data.player.choice_U)
    player_round_mat_to_sheet(wb.create_sheet(title='choice_R'), data.player.choice_R)
    player_round_mat_to_sheet(wb.create_sheet(title='choice_D'), data.player.choice_D)
    player_round_mat_to_sheet(wb.create_sheet(title='choice_nei_L'), data.player.choice_nei_L)
    player_round_mat_to_sheet(wb.create_sheet(title='choice_nei_U'), data.player.choice_nei_U)
    player_round_mat_to_sheet(wb.create_sheet(title='choice_nei_R'), data.player.choice_nei_R)
    player_round_mat_to_sheet(wb.create_sheet(title='choice_nei_D'), data.player.choice_nei_D)
    wb.save(file_name)


def save_strategy_response(session_code, data: OTreeSessionData, w, h):
    """Save a excel with strategy response."""

    file_name = join(OUTPUT_DIR, f'{APP_NAME}_response_{session_code}.xlsx')
    wb = Workbook()

    ws0 = wb.active
    ws0.title = "All"
    ws0.append('Round player info '
               'choice_L info_L nei_L_last_choice '
               'choice_U info_U nei_U_last_choice '
               'choice_R info_R nei_R_last_choice '
               'choice_D info_D nei_D_last_choice'.split(' '))

    rnd_stats = stats_strategy_response(data, ws0)
    ws1 = wb.create_sheet(title='Stats')
    ws1.append('Round CC CD DC DD'.split() +
               list(map(lambda x: x[1] + x[0], product('CC CD DC DD'.split(), 'aa ah ha hh'.split()))))
    for rnd, stats in rnd_stats.items():
        ws1.append([rnd] + list(map(lambda s: sum(s.values()), stats.values())) +
                   sum(map(lambda s: list(s.values()), stats.values()), []))
    print('Done')

    wb.save(file_name)


def preprocess_data(data: OTreeSessionData):
    all_info = data.loc[:, ('player', 'info')] > 0
    data.loc[:, ('player', 'info')] = np.where(all_info, 'All', 'Highest')
    return data[data.player.choice_L.notna()]


def calculate_data(data: OTreeSessionData, w, h):
    nei = grid.neighbor_mapping(w, h)
    data = data.reindex(columns=data.columns.tolist()
                                + list(map(lambda x: ('player', x),
                                           'pid_L pid_U pid_R pid_D '
                                           'info_L info_U info_R info_D '
                                           'choice_nei_L choice_nei_U choice_nei_R choice_nei_D'.split(' '))))
    for idx, row in data.iterrows():
        rnd, pid = idx
        print(f'\rProducing neighbor info (round {rnd}, player {pid}) ...    ', end='')
        data.loc[idx, ('player', 'pid_L')] = nei[pid]['L']
        data.loc[idx, ('player', 'pid_U')] = nei[pid]['U']
        data.loc[idx, ('player', 'pid_R')] = nei[pid]['R']
        data.loc[idx, ('player', 'pid_D')] = nei[pid]['D']
        data.loc[idx, ('player', 'info_L')] = data.loc[(rnd, nei[pid]['L']), ('player', 'info')]
        data.loc[idx, ('player', 'info_U')] = data.loc[(rnd, nei[pid]['U']), ('player', 'info')]
        data.loc[idx, ('player', 'info_R')] = data.loc[(rnd, nei[pid]['R']), ('player', 'info')]
        data.loc[idx, ('player', 'info_D')] = data.loc[(rnd, nei[pid]['D']), ('player', 'info')]
        data.loc[idx, ('player', 'choice_nei_L')] = data.loc[(rnd, nei[pid]['L']), ('player', 'choice_R')]
        data.loc[idx, ('player', 'choice_nei_U')] = data.loc[(rnd, nei[pid]['U']), ('player', 'choice_D')]
        data.loc[idx, ('player', 'choice_nei_R')] = data.loc[(rnd, nei[pid]['R']), ('player', 'choice_L')]
        data.loc[idx, ('player', 'choice_nei_D')] = data.loc[(rnd, nei[pid]['D']), ('player', 'choice_U')]
    print('Done')
    return data


OPERATION = [
    save_neighbor_info,
    save_strategy_response,
]


def run():
    init(DATA_DIR, OUTPUT_DIR)
    file_name = select_data_file(DATA_DIR)
    session_code, data = load_session(file_name)
    data = data.rename(
        columns=dict(is_show_all_info='info',
                     choose_L='choice_L', choose_U='choice_U', choose_R='choice_R', choose_D='choice_D'))
    num_players = len(data.iloc[data.index.get_level_values('round_number') == 1])
    w, h = grid.ask_W_H(num_players) if not DEBUG else (DFT_W, DFT_H)
    data = preprocess_data(data)
    data = calculate_data(data, w, h)
    if not DEBUG:
        while True:
            OPERATION[option_choice(list(map(lambda x: x.__doc__, OPERATION)), "Choose one operation:")](
                session_code, data, w, h)
    else:
        # save_neighbor_info(session_code, data, w, h)
        # save_strategies_response(session_code, data, w, h)
        save_strategy_response_barplot(session_code, data, w, h)
        pass


if __name__ == '__main__':
    try:
        run()
    except Exception:
        traceback.print_exception(*sys.exc_info())
        if not DEBUG:
            os.system("pause")
